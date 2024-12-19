#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Oct 26 14:28:03 2023

@author: colinbehr
"""
import os
os.chdir('/Users/colinbehr/Desktop/RESEARCH')

import openpyxl

# Load the workbook and sheets
wb = openpyxl.load_workbook('ALL Bears Boxscore Data.xlsx')
boxscore_data = wb['Boxscore Data']
hourly_weather_data = wb['Hourly Weather Data']

# Counter for GameID to be added every 3 timestamps
game_id_counter = 1

# Populate the Hourly Weather Data
for row in range(2, boxscore_data.max_row + 1):
    # Get the timestamps for this row
    timestamp1 = boxscore_data.cell(row=row, column=boxscore_data.max_column - 2).value
    timestamp2 = boxscore_data.cell(row=row, column=boxscore_data.max_column - 1).value
    timestamp3 = boxscore_data.cell(row=row, column=boxscore_data.max_column).value
    
    timestamps = [timestamp1, timestamp2, timestamp3]
    
    for ts in timestamps:
        # Create a new row in Hourly Weather Data
        new_row = hourly_weather_data.max_row + 1
        hourly_weather_data.cell(row=new_row, column=1).value = game_id_counter
        hourly_weather_data.cell(row=new_row, column=2).value = ts

    # Increment game_id_counter every 3 timestamps
    game_id_counter += 1

# Save the changes in a new workbook
wb.save('ALL Bears Boxscore Data_new.xlsx')
wb.close()



