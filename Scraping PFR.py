#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Sep 17 18:40:14 2023

@author: colinbehr
"""
import os
os.chdir('/Users/colinbehr/Downloads') #change wd to downloads

import requests
from openpyxl import Workbook
import datetime
from lxml import html
import re
from time import sleep
from random import randint  

def main():
    print("Starting main function #BearDown...")
    years = list(range(2013, 2023))
    url = "https://www.pro-football-reference.com/teams/chi/{year}.htm"

    wb = Workbook()
    ws = wb.active
    ws.title = "Chicago Bears Boxscore Data"
    ws.append(["Team", "Bears Score", "Location", "Opponent", "Opponent Score", "Combined Score", "Date", "Time", "Weather", "Surface", "URL"])
    #add_stadium_locations_to_wb(wb, 'NFL Stadium Latitude and Longtitude.csv')

    #pbp ws
    ws_play_by_play = wb.create_sheet("Play-By-Play Data")
    ws_play_by_play.append(["SeasonID", "GameID", "PlayID", "Quarter", "Time", "Down", "ToGo", "Location", "Away Score", "Home Score", "Detail"])

    for year in years:
        print(f"Processing year {year}")
        scrape_season_data(url.format(year=year), ws, ws_play_by_play, year)
        sleep(randint(5, 10))  #sleep for 5-10 seconds to be a polite guest :)
        
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    wb.save(f'Chicago_Bears_Boxscore_Data_{timestamp_str}.xlsx')
    print("Main function completed.")

#def add_stadium_locations_to_wb(wb, csv_filename): #put stadium locations in excel file
#    with open(csv_filename, 'r', encoding='utf-8-sig') as file:
#        reader = csv.reader(file)
#        ws_stadium_locations = wb.create_sheet("NFL Stadium Locations")
#        
#        for row in reader:
#            ws_stadium_locations.append(row)

#def get_stadium_coords(team=None, wb=None):
#    ws_stadium_locations = wb['NFL Stadium Locations']
#    if team is not None:
#        for row in ws_stadium_locations.iter_rows(min_row=2, values_only=True):  # skip header
#            if team == row[0]:  # assuming team is the first column, latitude is 2nd and longitude is 3rd.
#                return row[1], row[2]
#        return None, None
#    else:
#        stadium_locations = {}
#        for row in ws_stadium_locations.iter_rows(min_row=2, values_only=True):
#           stadium_locations[row[0]] = {"latitude": row[1], "longitude": row[2]}
#        return stadium_locations

game_id = 0

def scrape_season_data(url, ws, ws_play_by_play, year):
    global game_id  #globalize to keep it incrementing & not restart each season

    print(f"Scraping season data from {url}")
    response = requests.get(url)
    print(f"HTTP Status Code for Season: {response.status_code}")

    if response.status_code != 200:
        print("Failed to scrape season data. Skipping this season.")
        return

    tree = html.fromstring(response.content)
    games = tree.xpath('//table[@id="games"]/tbody/tr')
    for game in games:
        boxscore_url = game.xpath('.//td[@data-stat="boxscore_word"]/a/@href')
        if boxscore_url:
            game_id += 1  # increment only if a valid boxscore_url is found
            boxscore_url = "https://www.pro-football-reference.com" + boxscore_url[0]
            scrape_boxscore(boxscore_url, ws)  # call boxscore
            
            # call pbp function
            scrape_pbp(boxscore_url, ws_play_by_play, year, game_id)
        sleep(randint(5, 10))  # sleep for 5-10 seconds to be a polite guest :)


    
def scrape_boxscore(url, ws, team_of_interest="Chicago Bears"):
    print(f"Scraping Game Data from {url}")
    try:
        response = requests.get(url)
        
        if response.status_code == 200:
            tree = html.fromstring(response.content)

            meta_description_list = tree.xpath('//meta[@property="og:description"]/@content')
            if meta_description_list:
                meta_description = meta_description_list[0]
                print(f"Scraped Meta Description: {meta_description}")

                match = re.search(r'([a-zA-Z\s]+) (\d+) at ([a-zA-Z\s]+) (\d+)', meta_description)

                if match:
                    away_team = match.group(1)
                    away_score = int(match.group(2))
                    home_team = match.group(3)
                    home_score = int(match.group(4))

                    if home_team != team_of_interest:
                        opponent_team = home_team
                        opponent_score = home_score
                        home_team = team_of_interest
                        home_score = away_score
                        location = "@"
                    else:
                        opponent_team = away_team
                        opponent_score = away_score
                        location = "vs"
                        
#                    if location == "@":
#                        latitude, longitude = get_stadium_coords(opponent_team, ws)
#                    else:
#                        latitude, longitude = get_stadium_coords(home_team, ws)

                    # surface type (in comment)
                    comments = tree.xpath('//comment()')
                    surface = "Unknown"
                    for comment in comments:
                        comment_tree = html.fromstring(comment.text)
                        surface_list = comment_tree.xpath('//tr[th[contains(text(), "Surface")]]/td/text()')
                        if surface_list:
                            surface = surface_list[0]
                            break
                    # get weather from comment block as well
                    weather = "Unknown"
                    for comment in comments:
                        comment_tree = html.fromstring(comment.text)
                        weather_list = comment_tree.xpath('//tr[th[contains(text(), "Weather")]]/td/text()')
                        if weather_list:
                            weather = weather_list[0]
                            break
                        
                    # date & time
                    date = tree.xpath('//div[@class="scorebox_meta"]/div[1]/text()')[0]
                    
                    time = tree.xpath('//div[@class="scorebox_meta"]/div[contains(strong/text(), "Start Time")]/text()')[0]
                    #from dateutil.parser import parse
                    #time_obj = parse(raw_time)
                    #time_24hr = time_obj.strftime('%H:%M')
                    #time = time_24hr
                    
                    # Add game data to games list
#                    game_data = {
#                        "latitude": latitude,
#                        "longitude": longitude,
#                        "timestamp": f"{date}T{time}"
#                    }
#                    games.append(game_data)
                    
                    #make API call
#                    if latitude and longitude:
#                        api_weather_data = get_weather_data(latitude, longitude, game_data["timestamp"])
#                        if api_weather_data:
#                            weather = api_weather_data.get("description", "Unknown")
                            
                    combined_score = home_score + opponent_score
                    ws.append([home_team, home_score, location, opponent_team, opponent_score, combined_score, date, time, weather, surface, url])

                    print(f"Team: {home_team}, Bears Score: {home_score}, Location: {location}, Opponent: {opponent_team}, Opponent Score: {opponent_score}, Combined Score: {combined_score}, Date: {date}, Time: {time}, Weather: {weather} Surface: {surface}, URL: {url}")

                else:
                    print("Scores could not be extracted. Regex did not match.")
            else:
                print("Could not find meta description. XPath query returned an empty list.")
        else:
            print("Failed to get boxscore. Skipping this game.")
            
    except Exception as e:
        print(f"An exception occurred: {e}")

def scrape_pbp(url, ws_play_by_play, season_id, game_id):
    print(f"Scraping play-by-play data from {url}")
    print('----------------------------------------------------------------------------------')
    try:
        #clean out comment
        response = requests.get(url)
        tree = html.fromstring(response.content)
        comment_table = tree.xpath('//div[@id="all_pbp"]//comment()')[0]
        comment_table_clean = comment_table.text.replace('<!--', '')
        comment_table_clean = comment_table_clean.replace('-->', '')
        tree_clean = html.fromstring(comment_table_clean)
        
        play_by_play_table = tree_clean.xpath('//table[@id="pbp"]/tbody/tr')
        play_id = 0  #play ID counter
        incremented_play_id = play_id #set new increment id to establish if not a play in ws
        
        for row in play_by_play_table:
            
            #scrape data for each column based on the HTML structure
            quarter = row.xpath('./th[@data-stat="quarter"]/text()')[0] if row.xpath('./th[@data-stat="quarter"]/text()') else ''
            time = row.xpath('string(./td[@data-stat="qtr_time_remain"])').strip()
            down = row.xpath('./td[@data-stat="down"]/text()')[0] if row.xpath('./td[@data-stat="down"]/text()') else ''
            to_go = row.xpath('./td[@data-stat="yds_to_go"]/text()')[0] if row.xpath('./td[@data-stat="yds_to_go"]/text()') else ''
            location = row.xpath('./td[@data-stat="location"]/text()')[0] if row.xpath('./td[@data-stat="location"]/text()') else ''
            away_score = row.xpath('./td[@data-stat="pbp_score_aw"]/text()')[0] if row.xpath('./td[@data-stat="pbp_score_aw"]/text()') else ''
            home_score = row.xpath('./td[@data-stat="pbp_score_hm"]/text()')[0] if row.xpath('./td[@data-stat="pbp_score_hm"]/text()') else ''
            detail = ''.join(row.xpath('./td[@data-stat="detail"]//text()')).strip()
            
            
            #check if all of the above values contain some data
            if all([quarter, time, down, to_go, location, away_score, home_score, detail]):
                incremented_play_id += 1  # increment playID
                curr_play_id = incremented_play_id #new playID if play exists
            else:
                curr_play_id = None #wasnt a play

            ws_play_by_play.append([season_id, game_id, curr_play_id if curr_play_id is not None else '', quarter, time, down, to_go, location, away_score, home_score, detail])
            
    except Exception as e:
          print(f"An exception occurred: {e}")


#start er up
if __name__ == "__main__":
    main()
