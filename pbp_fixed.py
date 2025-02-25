#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Dec 11 11:43:37 2023

@author: colinbehr
"""

import requests
from lxml import html
from openpyxl import load_workbook
import time
import random

import os
os.chdir('/Users/colinbehr/Desktop/RESEARCH')

# Function to scrape play-by-play data
def scrape_pbp(url, play_by_play_sheet, season_id, game_id):
    print(f"Scraping play-by-play data from {url}")
    print('----------------------------------------------------------------------------------')
    try:
        # Clean out comment
        response = requests.get(url)
        tree = html.fromstring(response.content)
        comment_table = tree.xpath('//div[@id="all_pbp"]//comment()')[0]
        comment_table_clean = comment_table.text.replace('<!--', '').replace('-->', '')
        tree_clean = html.fromstring(comment_table_clean)
        
        play_by_play_table = tree_clean.xpath('//table[@id="pbp"]/tbody/tr')
        play_id = 0  # Play ID counter
        
        for row in play_by_play_table:
            # Scrape data for each column based on the HTML structure
            quarter = row.xpath('./th[@data-stat="quarter"]/text()')[0] if row.xpath('./th[@data-stat="quarter"]/text()') else ''
            time = row.xpath('string(./td[@data-stat="qtr_time_remain"])').strip()
            down = row.xpath('./td[@data-stat="down"]/text()')[0] if row.xpath('./td[@data-stat="down"]/text()') else ''
            to_go = row.xpath('./td[@data-stat="yds_to_go"]/text()')[0] if row.xpath('./td[@data-stat="yds_to_go"]/text()') else ''
            location = row.xpath('./td[@data-stat="location"]/text()')[0] if row.xpath('./td[@data-stat="location"]/text()') else ''
            away_score = row.xpath('./td[@data-stat="pbp_score_aw"]/text()')[0] if row.xpath('./td[@data-stat="pbp_score_aw"]/text()') else ''
            home_score = row.xpath('./td[@data-stat="pbp_score_hm"]/text()')[0] if row.xpath('./td[@data-stat="pbp_score_hm"]/text()') else ''
            detail = ''.join(row.xpath('./td[@data-stat="detail"]//text()')).strip()
            
            # Check if all of the above values contain some data
            if all([quarter, time, down, to_go, location, away_score, home_score, detail]):
                play_id += 1  # Increment playID if it's a valid play
                curr_play_id = play_id  # Set current play ID
            else:
                curr_play_id = None  # It wasn't a valid play
                
            # Append data to the play_by_play_sheet
            play_by_play_sheet.append([
                season_id, game_id, curr_play_id if curr_play_id is not None else '',
                quarter, time, down, to_go, location, away_score, home_score, detail
            ])
            
    except Exception as e:
        print(f"An exception occurred: {e}")


# Load the workbook and select sheets
workbook = load_workbook(filename='ALL Bears Data.xlsx')
boxscore_data_sheet = workbook['Boxscore']  # Adjust sheet name as necessary
play_by_play_sheet = workbook.create_sheet('Play-By-Play')  # This will create a new sheet

# Set up headers for the Play-By-Play sheet
play_by_play_headers = [
    "SeasonID", "GameID", "PlayID", "Quarter", "Time",
    "Down", "ToGo", "Location", "Away Score", "Home Score", "Detail"
]
play_by_play_sheet.append(play_by_play_headers)

url_column_index = None
for cell in boxscore_data_sheet[1]:  # Assuming the first row contains headers
    if cell.value == "URL":
        url_column_index = cell.column  # Get the column index
        break

if url_column_index is None:
    raise ValueError("The 'URL' column was not found in the 'Boxscore' sheet.")

game_id = 1  # Initialize GameID
for row in boxscore_data_sheet.iter_rows(min_row=2, values_only=True):
    url = row[url_column_index - 1]  # Adjust for zero-based index

    # Verify that we have a proper URL
    if isinstance(url, str) and (url.startswith('http://') or url.startswith('https://')):
        season_id = url.split('/')[-1][:4]
        scrape_pbp(url, play_by_play_sheet, season_id, game_id)
        game_id += 1  # Increment GameID for each new game
        
        # Sleep for a random time between 5 and 10 seconds
        time.sleep(random.randint(5, 10))
    else:
        print(f"Invalid URL found: {url}, skipping to the next one.")

# Save the workbook when done
workbook.save(filename='pbp updated.xlsx')

print("Scraping completed and data saved.")
