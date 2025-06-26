#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Oct 25 16:17:42 2023

@author: colinbehr
"""
import openpyxl
import requests
import os
from time import sleep
from random import randint

#set api request 
API_BASE_URL = "https://weather.visualcrossing.com/VisualCrossingWebServices/rest/services/timeline/"
API_KEY = "YOUR API KEY"

os.chdir('/Users/colinbehr/Desktop/RESEARCH') #wd

#create sheets
wb = openpyxl.load_workbook('ALL Bears Data_fix.xlsx')
input_sheet = wb['Game + Timestamp + Coords']
output_sheet = wb.create_sheet('HOURLY Weather 2')

#set in order of json ouput 
headers = ['GameID', 'Timestamp', 'Latitude', 'Longitude', 'datetime', 'datetimeEpoch', 'temp', 'feelslike', 'humidity', 
           'dew', 'precip', 'precipprob', 'snow', 'snowdepth', 'preciptype', 'windgust', 'windspeed', 'winddir', 
           'pressure', 'visibility', 'cloudcover', 'solarradiation', 'solarenergy', 'uvindex', 'conditions', 'icon', 'sunset']
output_sheet.append(headers)

def make_api_call(game_id, latitude, longitude, timestamp):
    url = f"{API_BASE_URL}{latitude},{longitude}/{timestamp}?key={API_KEY}&contentType=json&include=current"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if 'currentConditions' in data and len(data['currentConditions']) > 0:
            day_data = data['currentConditions']
            #convert all lists to strings prior to appending to ouput
            row_data = [game_id, timestamp, latitude, longitude] + [str(day_data.get(header, None)) for header in headers[4:]]
            output_sheet.append(row_data)
        else:
            print(f"No 'CC' field for timestamp {timestamp}")
    else:
        print(f"Error for timestamp {timestamp} with status code: {response.status_code}")
    #random sleep to avoid overloading requests
    sleep(randint(5, 10))

# Iterate over every row in the input sheet and make API calls
for row in input_sheet.iter_rows(min_row=2, values_only=True):  
    game_id, timestamp, latitude, longitude = row
    make_api_call(game_id, latitude, longitude, timestamp)
    #if "2013" in timestamp: #uncomment if only want one year 
    #    make_api_call(game_id, latitude, longitude, timestamp)

# Save the workbook with the new data
wb.save('FULL Weather Data.xlsx')
wb.close()
