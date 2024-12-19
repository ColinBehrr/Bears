#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Oct 26 14:56:45 2023

@author: colinbehr
"""
import os
os.chdir('/Users/colinbehr/Desktop/RESEARCH')

import openpyxl
import datetime

# Load the workbook and sheets
wb = openpyxl.load_workbook('ALL Bears Boxscore Data.xlsx')
boxscore_data = wb['Boxscore Data']
stadium_locations = wb['Stadium Locations']
hourly_weather_data = wb['Hourly Weather Data']

# Get column indices by name
boxscore_location_col = [cell.value for cell in boxscore_data[1]].index("Location")
boxscore_gameid_col = [cell.value for cell in boxscore_data[1]].index("GameID")
boxscore_opponent_col = [cell.value for cell in boxscore_data[1]].index("Opponent")
boxscore_date_col = [cell.value for cell in boxscore_data[1]].index("Date")
stadium_team_col = [cell.value for cell in stadium_locations[1]].index("Team")
stadium_lat_col = [cell.value for cell in stadium_locations[1]].index("Latitude")
stadium_long_col = [cell.value for cell in stadium_locations[1]].index("Longitude")
hourly_gameid_col = [cell.value for cell in hourly_weather_data[1]].index("GameID")
hourly_lat_col = [cell.value for cell in hourly_weather_data[1]].index("Latitude")
hourly_long_col = [cell.value for cell in hourly_weather_data[1]].index("Longitude")

# Extract latitude and longitude for "Chicago Bears"
for row in stadium_locations.iter_rows(min_row=2):  # Start from second row to skip headers
    if row[stadium_team_col].value == "Chicago Bears":
        bears_lat = row[stadium_lat_col].value
        bears_long = row[stadium_long_col].value
        break

# Create a list of GameIDs that are marked as "vs"
vs_game_ids = []
for row in boxscore_data.iter_rows(min_row=2):  # Skipping header row
    if row[boxscore_location_col].value == "vs":  
        vs_game_ids.append(row[boxscore_gameid_col].value)  

# Update "Hourly Weather Data" for matching GameIDs
for row in hourly_weather_data.iter_rows(min_row=2):  # Skipping header row
    if row[hourly_gameid_col].value in vs_game_ids:  # Check if GameID matches
        row[hourly_lat_col].value = bears_lat  # Update latitude
        row[hourly_long_col].value = bears_long  # Update longitude

# Create a dictionary to map teams to their latitude and longitude
team_location_dict = {}
for row in stadium_locations.iter_rows(min_row=2):
    team_location_dict[row[stadium_team_col].value] = (row[stadium_lat_col].value, row[stadium_long_col].value)

# Create a list of GameIDs that are marked as "@"
at_game_ids = {}
for row in boxscore_data.iter_rows(min_row=2):
    if row[boxscore_location_col].value == "@":
        opponent = row[boxscore_opponent_col].value
        game_date = datetime.datetime.strptime(row[boxscore_date_col].value, "%A %b %d, %Y")
        if opponent == "Atlanta Falcons" and game_date < datetime.datetime(2017, 8, 1):
            opponent = "Atlanta Falcons (Georgia Dome)"
        elif opponent == "Atlanta Falcons" and game_date >= datetime.datetime(2017, 8, 1):
            opponent = "Atlanta Falcons (Mercedes Benz)"
        at_game_ids[row[boxscore_gameid_col].value] = team_location_dict[opponent]  # Store the opponent's lat and long

# Update "Hourly Weather Data" for matching GameIDs
for row in hourly_weather_data.iter_rows(min_row=2):
    game_id = row[hourly_gameid_col].value
    if game_id in at_game_ids:
        row[hourly_lat_col].value, row[hourly_long_col].value = at_game_ids[game_id]

# Save changes to the workbook
wb.save('ALL Bears Boxscore Data_updated.xlsx')
wb.close()






